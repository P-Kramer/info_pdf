def checar_divergencias(df_at, df_cd):
    import pandas as pd
    import numpy as np
    from rapidfuzz import fuzz, process
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    import re
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter
    # Configuração de cor
    YELLOW_FILL = PatternFill(start_color='FFFFF000', end_color='FFFFF000', fill_type='solid')
    HEADER_FILL = PatternFill(start_color='FFDDDDDD', end_color='FFDDDDDD', fill_type='solid')
    HEADER_FONT = Font(bold=True)
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
    # Mapeamento manual de pareamentos forçados
    pareamentos_forcados = {
        "AMERCO /NV/":"UHAL'B",
        "POLEN CAPITAL FOCUS US GR USD INSTL":"PFUGI",
        "JPM US SELECT EQUITY PLUS C (ACC) USD":"SELQZ",
        "AMUNDI FDS US EQ FUNDM GR I2 USD C":"PONVS",
        "MS INVF GLOBAL ENDURANCE I USD ACC":"SMFVZ",
        "PRINCIPAL PREFERRED SECS N INC USD": "PRGPZ",
        "PIMCO GIS INCOME H INSTL USD INC": "PCOAZ"
    }
    def formatar_aba(ws, colunas_monetarias=None, colunas_percentuais=None):
        colunas_monetarias = colunas_monetarias or []
        colunas_percentuais = colunas_percentuais or []
        
        # Cabeçalhos
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN

        # Linhas de dados
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.alignment = CENTER_ALIGN

        # Formatar colunas numéricas específicas
        for col in colunas_monetarias:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'

        for col in colunas_percentuais:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = '0.00%'

        # Autoajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
    # Função para extrair CUSIP do texto
    def extrair_cusip(texto):
        match = re.search(r'US([A-Z0-9]{9})', str(texto).upper())
        return match.group(1) if match else None

    # Pré-processamento
    def processar_com_dinheiro(df):
        df = df[
            (df['Carteira'].str.startswith('LA_', na=False)) &
            (df['Classe'].isin(['EQUITY', "FIXED INCOME","FLOATING INCOME"]))
        ].copy()
        df = df.rename(columns={
            'Ativo': 'Ticker',
            'Quant.': 'Quantidade',
            'Saldo Bruto': 'MarketValue'
        })
        df['TickerBase'] = df['ticker_cmd_puro'].str.split(':').str[-1].str.strip()
        return df[['Descrição', 'Ticker', 'TickerBase', 'Quantidade', 'MarketValue', 'Classe']]

    def processar_ativos(df):
        df = df.rename(columns={
            'Ativo': 'Nome',
            'Quantidade Total': 'Quantidade',
            'Market Value': 'MarketValue'
        })
        df['TickerBase'] = df['Ticker'].str.extract(r'([A-Z]{2,6}$)')[0].fillna(df['Ticker'])
        return df[['Nome', 'Ticker', 'TickerBase', 'Quantidade', 'MarketValue','CUSIP']]


    equity_cd = processar_com_dinheiro(df_cd)
    equity_at = processar_ativos(df_at)

    # Mapeamento forçado: {CUSIP: Descrição}
    pareamentos_cusip_descricao_forcados = {
        "J7596PAJ8": "SOFTBANK GROUP 17/UND. 6,875%",
    }

    # === PAREAMENTO FORÇADO POR CUSIP E DESCRIÇÃO ===
    forced_cusip_desc_matches = []

    # Percorrer os ATIVOS
    for _, row_at in equity_at.iterrows():
        cusip_at = row_at['CUSIP']
        
        if pd.notna(cusip_at) and cusip_at in pareamentos_cusip_descricao_forcados:
            descricao_alvo = pareamentos_cusip_descricao_forcados[cusip_at]
            
            # Buscar no COM DINHEIRO pela descrição alvo
            match_cd = equity_cd[equity_cd['Descrição'].str.contains(descricao_alvo, case=False, na=False)]
            
            if not match_cd.empty:
                row_cd = match_cd.iloc[0]
                forced_cusip_desc_matches.append({
                    'Descrição_CD': row_cd['Descrição'],
                    'Ticker_CD': row_cd['Ticker'],
                    'TickerBase': row_cd['TickerBase'],
                    'Classe': row_cd["Classe"],
                    'Quantidade_CD': row_cd['Quantidade'],
                    'MarketValue_CD': row_cd['MarketValue'],
                    'Nome_MS': row_at['Nome'],
                    'Ticker_MS': row_at['Ticker'],
                    'Quantidade_MS': row_at['Quantidade'],
                    'CUSIP_MS': row_at['CUSIP'],
                    'MarketValue_MS': row_at['MarketValue'],
                    'Similaridade': 100
                })
                
                # Remover os itens pareados

                equity_cd = equity_cd.drop(row_cd.name)
                equity_at = equity_at.drop(row_at.name)
                
    # Converter para DataFrame
    forced_cusip_desc_df = pd.DataFrame(forced_cusip_desc_matches)


    # === PAREAMENTO POR CUSIP ===
    equity_cd['CUSIP_EXTRAIDO'] = equity_cd['Ticker'].apply(extrair_cusip)
    equity_at['CUSIP'].replace('', pd.NA, inplace=True)
    equity_at['CUSIP'].replace(' ', pd.NA, inplace=True)
    equity_at['CUSIP'].replace(' ', pd.NA, inplace=True) 
    cd_com_cusip = equity_cd.dropna(subset=['CUSIP_EXTRAIDO'])
    at_com_cusip = equity_at.dropna(subset=['CUSIP'])

    cd_sem_cusip = equity_cd[equity_cd['CUSIP_EXTRAIDO'].isna()]
    at_sem_cusip = equity_at[equity_at['CUSIP'].isna()]

    cusip_matches = []
    for _, row_cd in cd_com_cusip.iterrows():
        for _, row_at in at_com_cusip.iterrows():
            if row_at['CUSIP'] in row_cd['Ticker']:
                cusip_matches.append({
                    'Descrição_CD': row_cd['Descrição'],
                    'Ticker_CD': row_cd['Ticker'],
                    'TickerBase': row_cd['TickerBase'],
                    'Classe': row_cd["Classe"],
                    'Quantidade_CD': row_cd['Quantidade'],
                    'MarketValue_CD': row_cd['MarketValue'],
                    'Nome_MS': row_at['Nome'],
                    'Ticker_MS': row_at['Ticker'],
                    'Quantidade_MS': row_at['Quantidade'],
                    'CUSIP_MS': row_at['CUSIP'],
                    'MarketValue_MS': row_at['MarketValue'],
                    'Similaridade': 100
                })
                break

    cusip_df = pd.DataFrame(cusip_matches)

    # Atualiza para seguir só com os SEM CUSIP
    equity_cd = cd_sem_cusip
    equity_at = at_sem_cusip

    # Matching exato
    exact_match = pd.merge(
        equity_cd,
        equity_at,
        on='TickerBase',
        suffixes=('_CD', '_MS'),
        how='inner'
    )

    # Transformar em lista de dicionários no formato padrão
    exact_matches_list = []
    for index, row in exact_match.iterrows():
        exact_matches_list.append({
            'Descrição_CD': row['Descrição'],
            'Ticker_CD': row['Ticker_CD'],
            'TickerBase': row['TickerBase'],
            'Classe': row['Classe'],
            'Quantidade_CD': row['Quantidade_CD'],
            'MarketValue_CD': row['MarketValue_CD'],
            'Nome_MS': row['Nome'],
            'Ticker_MS': row['Ticker_MS'],
            'Quantidade_MS': row['Quantidade_MS'],
            'CUSIP_MS': row['CUSIP'],
            'MarketValue_MS': row['MarketValue_MS'],
            'Similaridade': 100
        })

    matched_tickers = exact_match['TickerBase'].unique()

    remaining_cd = equity_cd[~equity_cd['TickerBase'].isin(matched_tickers)]
    remaining_at = equity_at[~equity_at['TickerBase'].isin(matched_tickers)]

    # Matching forçado
    forcados = []
    for _, row in remaining_cd.iterrows():
        descricao_upper = row['Descrição'].strip().upper()
        if descricao_upper in pareamentos_forcados:
            ticker_alvo = pareamentos_forcados[descricao_upper]
            match_row = remaining_at[remaining_at['Ticker'] == ticker_alvo]
            if not match_row.empty:
                match_row = match_row.iloc[0]
                forcados.append({
                    'Descrição_CD': row['Descrição'],
                    'Ticker_CD': row['Ticker'],
                    'TickerBase': row['TickerBase'],
                    'Classe': row["Classe"],
                    'Quantidade_CD': row['Quantidade'],
                    'MarketValue_CD': row['MarketValue'],
                    'Nome_MS': match_row['Nome'],
                    'Ticker_MS': match_row['Ticker'],
                    'Quantidade_MS': match_row['Quantidade'],
                    'CUSIP_MS': match_row ['CUSIP'],
                    'MarketValue_MS': match_row['MarketValue'],
                    'Similaridade': 100
                })

    # Fuzzy por descrição completa
    fuzzy_matches = []
    for _, row in remaining_cd.iterrows():
        match, score, _ = process.extractOne(
            row['Descrição'],
            remaining_at['Nome'].tolist(),
            scorer=fuzz.token_set_ratio
        )
        if score >= 85:
            match_row = remaining_at[remaining_at['Nome'] == match].iloc[0]
            fuzzy_matches.append({
                'Descrição_CD': row['Descrição'],
                'Ticker_CD': row['Ticker'],
                'TickerBase': row['TickerBase'],
                'Classe': row["Classe"],
                'Quantidade_CD': row['Quantidade'],
                'MarketValue_CD': row['MarketValue'],
                'Nome_MS': match_row['Nome'],
                'Ticker_MS': match_row['Ticker'],
                'Quantidade_MS': match_row['Quantidade'],
                'CUSIP_MS': match_row ['CUSIP'],
                'MarketValue_MS': match_row['MarketValue'],
                'Similaridade': score
            })

    # Junta todos os pareamentos
    all_matches = pd.concat([
        forced_cusip_desc_df,  # Pareamentos forçados por CUSIP e descrição
        cusip_df,              # Pareamentos por CUSIP tradicional
        pd.DataFrame(exact_matches_list),
        pd.DataFrame(forcados + fuzzy_matches),
        # ... outros pareamentos
    ], ignore_index=True)

    # Outras rodadas de matching
    matched_bases = all_matches['TickerBase'].unique()
    extra_cd = equity_cd[~equity_cd['TickerBase'].isin(matched_bases)]
    extra_at = equity_at[~equity_at['TickerBase'].isin(matched_bases)]

    ticker_matches = []
    for _, row_cd in extra_cd.iterrows():
        match_row = extra_at[extra_at['Ticker'] == row_cd['TickerBase']]
        if not match_row.empty:
            match_row = match_row.iloc[0]
            ticker_matches.append({
                'Descrição_CD': row_cd['Descrição'],
                'Ticker_CD': row_cd['Ticker'],
                'TickerBase': row_cd['TickerBase'],
                'Classe': row_cd["Classe"],
                'Quantidade_CD': row_cd['Quantidade'],
                'MarketValue_CD': row_cd['MarketValue'],
                'Nome_MS': match_row['Nome'],
                'Ticker_MS': match_row['Ticker'],
                'Quantidade_MS': match_row['Quantidade'],
                'CUSIP_MS': match_row ['CUSIP'],
                'MarketValue_MS': match_row['MarketValue'],
                'Similaridade': 100
            })

    cd_restante = extra_cd[~extra_cd['TickerBase'].isin([m['TickerBase'] for m in ticker_matches])]
    at_restante = extra_at[~extra_at['TickerBase'].isin([m['TickerBase'] for m in ticker_matches])]

    palavra_matches = []
    for _, row_cd in cd_restante.iterrows():
        palavra_cd = str(row_cd['Descrição']).strip().split()[0].upper()
        candidatos = at_restante['Nome'].tolist()
        melhores = process.extract(palavra_cd, candidatos, scorer=fuzz.token_sort_ratio, limit=1)
        if melhores:
            match, score, _ = melhores[0]
            if score >= 80:
                match_row = at_restante[at_restante['Nome'] == match].iloc[0]
                palavra_matches.append({
                    'Descrição_CD': row_cd['Descrição'],
                    'Ticker_CD': row_cd['Ticker'],
                    'TickerBase': row_cd['TickerBase'],
                    'Classe': row_cd["Classe"],
                    'Quantidade_CD': row_cd['Quantidade'],
                    'MarketValue_CD': row_cd['MarketValue'],
                    'Nome_MS': match_row['Nome'],
                    'Ticker_MS': match_row['Ticker'],
                    'Quantidade_MS': match_row['Quantidade'],
                    'CUSIP_MS': match_row ['CUSIP'],
                    'MarketValue_MS': match_row['MarketValue'],
                    'Similaridade': score
                })

    extra_df = pd.DataFrame(ticker_matches + palavra_matches)
    all_matches = pd.concat([all_matches, extra_df], ignore_index=True)

    # Cálculos de diferença
    if not all_matches.empty:
        all_matches['PrecoUnitario_CD'] = all_matches['MarketValue_CD'] / all_matches['Quantidade_CD']
        all_matches['PrecoUnitario_MS'] = all_matches['MarketValue_MS'] / all_matches['Quantidade_MS']
        all_matches['Diff_Quantidade'] = all_matches['Quantidade_CD'] - all_matches['Quantidade_MS']
        all_matches['Diff_MarketValue'] = all_matches['MarketValue_CD'] - all_matches['MarketValue_MS']
        all_matches['Diff_PrecoUnitario'] = all_matches['PrecoUnitario_CD'] - all_matches['PrecoUnitario_MS']
        all_matches['Pct_Diff_Quantidade'] = all_matches['Diff_Quantidade'] / all_matches['Quantidade_MS']
        all_matches['Pct_Diff_MarketValue'] = all_matches['Diff_MarketValue'] / all_matches['MarketValue_MS']
        all_matches['Pct_Diff_PrecoUnitario'] = all_matches['Diff_PrecoUnitario'] / all_matches['PrecoUnitario_MS']
        all_matches['Destaque'] = np.abs(all_matches['Diff_PrecoUnitario']) > 1

    tickersbase_usados_cd = all_matches['TickerBase'].unique()
    tickers_usados_at = all_matches['Ticker_MS'].unique()

    non_matched_cd = equity_cd[~equity_cd['TickerBase'].isin(tickersbase_usados_cd)]
    non_matched_at = equity_at[~equity_at['Ticker'].isin(tickers_usados_at)]

    non_matched_consolidado = pd.concat([
        non_matched_cd.assign(Origem='COM DINHEIRO'),
        non_matched_at.assign(Origem='ATIVOS')
    ], ignore_index=True)

    # Ordenação final
    col_order = [
        'TickerBase', 'Ticker_MS', 'Ticker_CD', 'CUSIP_MS', 'Descrição_CD', 'Nome_MS',
        'Quantidade_CD', 'Quantidade_MS', 'Diff_Quantidade',
        'MarketValue_CD', 'MarketValue_MS', 'Diff_MarketValue', 'Pct_Diff_MarketValue', "Classe"
    ]
    all_matches = all_matches[col_order]

    # Exporta para Excel
    wb = Workbook()
    wb.remove(wb.active)

    ws_pareados = wb.create_sheet("Pareados")
    for r in dataframe_to_rows(all_matches, index=False, header=True):
        ws_pareados.append(r)

    ws_nao_pareados = wb.create_sheet("Não Pareados")
    for r in dataframe_to_rows(non_matched_consolidado, index=False, header=True):
        ws_nao_pareados.append(r)

    ws_so_cd = wb.create_sheet("Só COM DINHEIRO")
    for r in dataframe_to_rows(non_matched_cd, index=False, header=True):
        ws_so_cd.append(r)

    ws_so_at = wb.create_sheet("Só ATIVOS")
    for r in dataframe_to_rows(non_matched_at, index=False, header=True):
        ws_so_at.append(r)

    if not all_matches.empty:
        diff_mv_col = col_order.index('Diff_MarketValue') + 1
        pct_diff_mv_col = all_matches.columns.get_loc('Pct_Diff_MarketValue') + 1
        classe_col = all_matches.columns.get_loc('Classe') + 1

        for idx, row in enumerate(ws_pareados.iter_rows(min_row=2, max_row=ws_pareados.max_row), start=2):
            diff_mv = ws_pareados.cell(row=idx, column=diff_mv_col).value
            pct_diff_mv = ws_pareados.cell(row=idx, column=pct_diff_mv_col).value
            classe = ws_pareados.cell(row=idx, column=classe_col).value

            if classe == "EQUITY":
                destacar = diff_mv is not None and abs(diff_mv) > 1
            else:
                destacar = pct_diff_mv is not None and abs(pct_diff_mv) > 0.01

            if destacar:
                for cell in row:
                    cell.fill = YELLOW_FILL
    # Índices das colunas a formatar
    col_monetarias = [col_order.index(c) + 1 for c in ['MarketValue_CD', 'MarketValue_MS', 'Diff_MarketValue']]
    col_percentuais = [col_order.index('Pct_Diff_MarketValue') + 1]

    formatar_aba(ws_pareados, colunas_monetarias=col_monetarias, colunas_percentuais=col_percentuais)
    formatar_aba(ws_nao_pareados)
    formatar_aba(ws_so_cd)
    formatar_aba(ws_so_at)
    # Exporta em memória (para download via Streamlit)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return all_matches, buffer