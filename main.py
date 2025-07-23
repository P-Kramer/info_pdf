import fitz  # PyMuPDF
import pandas as pd
import re
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

def processar_pdf (pdf_bytes , return_excel=False):

    dados = []
    ativos_sem_total = []

    # Função para verificar se o span está em negrito
    def is_bold(span):
        return "Bold" in span.get("font", "")

    ativo_atual = None
    total_extraido = False
    linhas_ativas = []
    cusip = None

    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
        for i, page in enumerate(doc):
            spans = []
            blocks = page.get_text("dict")["blocks"]
            for block in blocks:
                for line in block.get("lines", []):
                    for span in line.get("spans", []):
                        spans.append(span)
            idx = 0
            while idx < len(spans):
                span = spans[idx]
                texto = span["text"].strip()

                if "CUSIP" in texto.upper():
                    match = re.search(r'CUSIP\s+([A-Z0-9]+)', texto.upper())
                    if match:
                        cusip = match.group(1)
                        for back in range(1, 10):
                            if idx - back >= 0:
                                span_back = spans[idx - back]
                                texto_back = span_back.get("text", "").strip()
                                if is_bold(span_back) and len(texto_back) > 6:
                                    ativo_atual = texto_back
                                    total_extraido = False
                                    linhas_ativas = []

                                    # === Busca valores para Face Value e Market Value ===
                                    face_value = None
                                    market_value = None

                                    for forward in range(1, 20):
                                        if idx + forward < len(spans):
                                            span_forward = spans[idx + forward]
                                            texto_forward = spans[idx + forward].get("text", "").strip()

                                            # Tenta identificar o Face Value (ex: 50,000.000)
                                            if face_value is None:
                                                match_face = re.match(r"^\d{1,3}(,\d{3})*(\.\d+)?$", texto_forward)
                                                if match_face:
                                                    try:
                                                        face_value = float(texto_forward.replace(",", ""))
                                                    except:
                                                        pass

                                            if market_value is None:
                                                if is_bold(span_forward):
                                                    texto_forward = span_forward.get("text", "").strip()
                                                    match_mv = re.match(r"^\$?\(?\d[\d,]*(\.\d{2})?\)?$", texto_forward)
                                                    if match_mv:
                                                        try:
                                                            market_value = float(
                                                                texto_forward.replace("$", "")
                                                                            .replace(",", "")
                                                                            .replace("(", "")
                                                                            .replace(")", "")
                                                            )
                                                        except:
                                                            pass

                                            if face_value and market_value:
                                                break

                                    ativos_sem_total.append({
                                        "Página": i + 1,
                                        "Ativo": ativo_atual,
                                        "CUSIP": cusip,
                                        "Face Value": face_value,
                                        "Market Value": market_value
                                    })
                                    break
                # Detecta novo ativo válido
                if is_bold(span) and "(" in texto and ")" in texto:
                    if any(palavra in texto for palavra in ["$", "/", "(MMF)", "(NL)", ",", "Approved List", "Focus List", "Investment Objectives", "Asset Class"]):
                        idx += 1
                        continue

                    # Se havia ativo anterior sem total, tenta pegar última linha válida com números
                    if ativo_atual and not total_extraido and linhas_ativas:
                        c= 0
                        for linha in (linhas_ativas):
                            textos = [span.get("text", "").strip() for span in linha if "text" in span]

                            while c == 0:
                                infos = textos
                                c+=1
                                try:
                                    
                                    quantidade = float(infos[1].replace(",", ""))
                                    total_cost = float(infos[4].replace(",", ""))
                                    market_value = float(infos[5].replace(",", ""))
                                    tickers = re.findall(r"\(([^)]+)\)", ativo_atual)
                                    ticker = tickers[-1].strip() if tickers else ""

                                    dados.append({
                                        "Página": i + 1,
                                        "Ativo": ativo_atual,
                                        "Ticker": ticker,
                                        "CUSIP": cusip if cusip else "",
                                        "Quantidade Total": total_cost if cusip else quantidade,
                                        "Total Cost": "" if cusip else total_cost,
                                        "Market Value": market_value
                                    })
                                    break
                                except Exception as e:
                                    print(f"Erro na extração alternativa de {ativo_atual}: {e}")

                    # Atualiza o novo ativo
                    ativo_atual = texto
                    total_extraido = False
                    linhas_ativas = []
                    idx += 1
                    continue

                if ativo_atual and re.search(r"\d", texto):
                    linha_nova = spans[idx:idx+6]
                    linhas_ativas.append(linha_nova)

                if ativo_atual and texto.startswith("Total"):
                    texto_total = ' '.join([sp["text"] for sp in spans[idx:idx+6]])
                    numeros = re.findall(r"[-]?[\d,.]+", texto_total)
                    if len(numeros) >= 3:
                        try:
                            quantidade = float(numeros[0].replace(",", ""))
                            total_cost = float(numeros[1].replace(",", ""))
                            if not cusip or ativo_atual == "AMUNDI PIO US EQ FNDM GW I2(C) (PONVS)":
                                market_value = float(numeros[2].replace(",", ""))
                            if cusip and ativo_atual != "AMUNDI PIO US EQ FNDM GW I2(C) (PONVS)" :
                                market_value = float(numeros[3].replace(",", ""))
                            tickers = re.findall(r"\(([^)]+)\)", ativo_atual)
                            ticker = tickers[-1].strip() if tickers else ""
                            if ativo_atual == "AMUNDI PIO US EQ FNDM GW I2(C) (PONVS)":
                                dados.append({
                                    "Página": i + 1,
                                    "Ativo": ativo_atual if ativo_atual else "Desconhecido",
                                    "Ticker": ticker,
                                    "CUSIP": "",
                                    "Quantidade Total": quantidade,
                                    "Total Cost": total_cost,
                                    "Market Value": market_value
                                })
                            else:             
                                dados.append({
                                    "Página": i + 1,
                                    "Ativo": ativo_atual if ativo_atual else "Desconhecido",
                                    "Ticker": ticker,
                                    "CUSIP": cusip if cusip else "",
                                    "Quantidade Total": quantidade,
                                    "Total Cost": None if cusip else total_cost,
                                    "Market Value": market_value
                                })
                            ativos_sem_total = [a for a in ativos_sem_total if a["CUSIP"] != cusip]
                            total_extraido = True
                            ativo_atual = None
                            linhas_ativas = []
                            cusip = None
                        except Exception as e:
                            print(f"Erro ao processar '{ativo_atual}': {e}")
                idx += 1

    # Junta os ativos com CUSIP mas sem total na lista principal
    for ativo in ativos_sem_total:
        dados.append({
            "Página": ativo["Página"],
            "Ativo": ativo["Ativo"],
            "Ticker": "",
            "CUSIP": ativo["CUSIP"],
            "Quantidade Total": ativo["Face Value"],
            "Total Cost": None,
            "Market Value": ativo["Market Value"]
        })

    # Cria e exporta o DataFrame para Excel em memória
    df = pd.DataFrame(dados)
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Ativos")
        workbook = writer.book
        worksheet = writer.sheets["Ativos"]

        border = Border(left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000'))
        header_font = Font(bold=True, color="000000")
        header_fill = PatternFill("solid", fgColor="FF3D98FF")

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = border
                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = header_fill

        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            worksheet.column_dimensions[col_letter].width = max_length + 2

    output.seek(0)

    print("✅ Extração e formatação concluídas.")
    if return_excel:
        return df, output
    else:
        return df, None
